"""Shared branded report builders (Excel + PDF) for the admin/CDSO reports.

Every report carries a text reconstruction of the Saint Louis College
letterhead — seal, "Saint Louis College", "of San Fernando, La Union", the
"Beacon of Wisdom" tagline, and the accreditation line — followed by the
brand-coloured table and a footer with generation stamp + page numbers.
"""
import os
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone as tz

REPORT_BRAND_HEX = '2A2B61'
REPORT_NAVY_HEX  = '1B2A63'
REPORT_LOGO_PATH = os.path.join(settings.BASE_DIR, 'report_assets', 'slclogo.jpg')

# Letterhead text (reconstruction of the official SLC letterhead).
LH_INSTITUTION = 'Saint Louis College'
LH_LOCATION    = 'of San Fernando, La Union'
LH_TAGLINE     = 'The Beacon of Wisdom in the North'
LH_ACCRED      = ('•  Center of Excellence in Teacher Education        '
                  '•  ISO 9001: 2015 Quality Management System Certified        '
                  '•  CHED Deregulated Status')


def report_filename(report_name, ext):
    """Filesystem-safe, human-readable report filename with date + time.

    e.g. report_filename('Audit Log Report', 'pdf')
         -> 'Audit Log Report - 2026-07-22 09-30 PM.pdf'
    """
    stamp = tz.localtime().strftime('%Y-%m-%d %I-%M %p')
    return f"{report_name} - {stamp}.{ext}"


# ── Letterhead fonts ────────────────────────────────────────────────────────
# "Saint Louis College" → Old English Text MT (blackletter); the tagline → a
# formal script. The PDF registers the Windows TTF and *embeds* the glyphs so
# the file renders correctly on any viewer. If the TTF is missing, the PDF
# falls back to Times. (Excel has no letterhead — it is a plain data table.)
_WIN_FONTS = os.path.join(os.environ.get('WINDIR', r'C:\Windows'), 'Fonts')
LH_INSTITUTION_TTF = 'OLDENGL.TTF'    # Old English Text MT

_fonts_ready = False
_PDF_INSTITUTION_FONT = 'Times-Bold'     # fallback until registered


def _register_letterhead_fonts():
    """Register the Old English TTF with reportlab once (embedded into each PDF).
    The tagline stays on the built-in Times-Italic."""
    global _fonts_ready, _PDF_INSTITUTION_FONT
    if _fonts_ready:
        return
    _fonts_ready = True
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    path = os.path.join(_WIN_FONTS, LH_INSTITUTION_TTF)
    if os.path.exists(path):
        try:
            pdfmetrics.registerFont(TTFont('SLC-OldEnglish', path))
            _PDF_INSTITUTION_FONT = 'SLC-OldEnglish'
        except Exception:
            pass


def branded_excel_response(*, filename, sheet_title, report_title, subtitle, headers, rows, col_widths):
    """Build a plain .xlsx data table — column headers on row 1, data from row 2.

    Deliberately has no letterhead/title band: merged cells and floating images
    break sorting, filtering and pivot tables. `report_title`/`subtitle` are
    accepted for API parity with the PDF builder but are not rendered.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    n = len(headers)
    last_col = chr(ord('A') + n - 1)

    # ── Column headers (row 1) ───────────────────────────────────────
    header_row = 1
    header_fill = PatternFill('solid', fgColor=REPORT_BRAND_HEX)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center')

    for i, width in enumerate(col_widths):
        ws.column_dimensions[chr(ord('A') + i)].width = width
    ws.freeze_panes = f'A{header_row + 1}'

    wrap = Alignment(wrap_text=True, vertical='top')
    r = header_row
    for row in rows:
        r += 1
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            if col == n:
                cell.alignment = wrap

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def draw_letterhead(canvas, w, h, *, title, footer_left='', footer_right=''):
    """Paint the SLC letterhead and footer onto `canvas` for a `w` x `h` page.

    Shared by the landscape report tables and the portrait registration
    confirmation, so the two can never drift apart. Everything is positioned
    relative to the page size, so it works in either orientation.
    """
    from reportlab.lib.units import mm
    from reportlab.lib import colors

    _register_letterhead_fonts()
    brand = colors.HexColor(f'#{REPORT_BRAND_HEX}')
    navy  = colors.HexColor(f'#{REPORT_NAVY_HEX}')

    canvas.saveState()
    left = 15 * mm
    cx = w / 2
    # Institution name is centred on the page; the seal tucks in just to its
    # left (a crest beside the name), vertically centred on the text block.
    inst_size = 22
    inst_w = canvas.stringWidth(LH_INSTITUTION, _PDF_INSTITUTION_FONT, inst_size)
    logo_w = 16 * mm
    if os.path.exists(REPORT_LOGO_PATH):
        try:
            logo_x = cx - inst_w / 2 - 6 * mm - logo_w
            canvas.drawImage(REPORT_LOGO_PATH, logo_x, h - 25 * mm,
                             width=logo_w, height=logo_w,
                             preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    canvas.setFillColor(navy)
    canvas.setFont(_PDF_INSTITUTION_FONT, inst_size)
    canvas.drawCentredString(cx, h - 15 * mm, LH_INSTITUTION)
    canvas.setFillColor(colors.HexColor('#333333'))
    canvas.setFont('Times-Roman', 10.5)
    canvas.drawCentredString(cx, h - 19.5 * mm, LH_LOCATION)
    canvas.setFillColor(colors.HexColor('#555555'))
    canvas.setFont('Times-Italic', 11)
    canvas.drawCentredString(cx, h - 24 * mm, LH_TAGLINE)
    # Accreditation line (centred)
    canvas.setFont('Times-Roman', 7.5)
    canvas.setFillColor(colors.HexColor('#555555'))
    canvas.drawCentredString(cx, h - 30 * mm, LH_ACCRED)
    # Rule
    canvas.setStrokeColor(navy)
    canvas.setLineWidth(1.2)
    canvas.line(left, h - 33 * mm, w - 15 * mm, h - 33 * mm)
    # Document title (centred)
    canvas.setFillColor(brand)
    canvas.setFont('Helvetica-Bold', 11)
    canvas.drawCentredString(cx, h - 39 * mm, title)
    # Footer
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#999999'))
    if footer_left:
        canvas.drawString(left, 10 * mm, footer_left)
    if footer_right:
        canvas.drawRightString(w - 15 * mm, 10 * mm, footer_right)
    canvas.restoreState()


def branded_pdf_response(*, filename, report_title, subtitle, generated_by, headers, rows,
                         col_widths_mm, extra_tables=()):
    """Build a landscape A4 PDF with the SLC letterhead, brand table and footer.

    `extra_tables` appends further titled tables below the first, each a dict of
    {title, headers, rows, col_widths_mm}. A report that counts the same rows
    along two different axes needs them kept apart: side by side in one row, the
    two sets of columns would each sum to the total, inviting a reader to add
    them all together and get twice the real figure.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    _register_letterhead_fonts()
    brand = colors.HexColor(f'#{REPORT_BRAND_HEX}')
    navy  = colors.HexColor(f'#{REPORT_NAVY_HEX}')
    generated_at = tz.localtime().strftime('%B %d, %Y %I:%M %p')

    def esc(s):
        return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def draw_frame(canvas, doc):
        w, h = landscape(A4)
        draw_letterhead(canvas, w, h, title=report_title,
                        footer_left=f'Generated {generated_at} by {generated_by} '
                                    f'· Saint Louis College CDSO · Confidential',
                        footer_right=f'Page {doc.page}')

    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        resp, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=44 * mm, bottomMargin=16 * mm, title=report_title,
    )
    cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=10)
    head_style = ParagraphStyle('head', fontName='Helvetica-Bold', fontSize=9,
                                textColor=colors.white, leading=11)
    sub_style = ParagraphStyle('sub', fontName='Helvetica', fontSize=9,
                               textColor=colors.HexColor('#666666'))

    section_style = ParagraphStyle('section', fontName='Helvetica-Bold', fontSize=10,
                                   textColor=navy, spaceBefore=0, spaceAfter=0)

    def build_table(t_headers, t_rows, t_widths):
        data = [[Paragraph(esc(h), head_style) for h in t_headers]]
        for row in t_rows:
            data.append([Paragraph(esc(c), cell_style) for c in row])
        if len(data) == 1:
            data.append([Paragraph('No records match the selected filters.', cell_style)]
                        + [Paragraph('', cell_style) for _ in range(len(t_headers) - 1)])

        table = Table(data, colWidths=[w * mm for w in t_widths], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F5FA')]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E8F0')),
        ]))
        return table

    story = [Paragraph(esc(subtitle), sub_style), Spacer(1, 4 * mm)]
    story.append(build_table(headers, rows, col_widths_mm))
    for extra in extra_tables:
        story.append(Spacer(1, 7 * mm))
        story.append(Paragraph(esc(extra['title']), section_style))
        story.append(Spacer(1, 3 * mm))
        story.append(build_table(extra['headers'], extra['rows'], extra['col_widths_mm']))
    doc.build(story, onFirstPage=draw_frame, onLaterPages=draw_frame)
    return resp
