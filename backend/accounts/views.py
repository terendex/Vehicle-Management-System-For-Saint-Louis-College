import os
from rest_framework import generics, permissions, status, filters
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils import timezone
from .models import User, AuditLog, Notification
from .serializers import (
    UserSerializer,
    UserUpdateSerializer,
    RegisterSerializer,
    AdminReplaceSerializer,
    GuardCreateSerializer,
    AdminOwnerCreateSerializer,
    CustomTokenObtainPairSerializer,
    AuditLogSerializer,
    NotificationSerializer,
)


def get_client_ip(request):
    """Extract client IP from request."""
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def log_action(request, action, target_user=None, details=''):
    """Create an audit log entry."""
    AuditLog.objects.create(
        actor=request.user,
        action=action,
        target_user=target_user,
        details=details,
        ip_address=get_client_ip(request),
    )


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class IsAdminRole(permissions.BasePermission):
    """Allow access only to users with admin role."""
    def has_permission(self, request, view):
        return (
            request.user
            and request.user.is_authenticated
            and request.user.role == 'admin'
        )


class IsAdminOrCdso(permissions.BasePermission):
    """Allow access to the CDSO (admin) role."""
    def has_permission(self, request, view):
        return (
            request.user
            and request.user.is_authenticated
            and request.user.role == 'admin'
        )


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class RegisterView(generics.CreateAPIView):
    queryset            = User.objects.all()
    serializer_class    = RegisterSerializer
    permission_classes  = [IsAdminRole]

    def perform_create(self, serializer):
        user = serializer.save()
        log_action(self.request, AuditLog.Action.USER_CREATED, target_user=user)


class MeView(generics.RetrieveAPIView):
    serializer_class   = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user


# ──────────────────────────────────────────────
#  User Management (admin only)
# ──────────────────────────────────────────────

class UserListView(generics.ListAPIView):
    """List all users except admins, with optional ?search= by name."""
    serializer_class   = UserSerializer
    permission_classes = [IsAdminRole]
    pagination_class   = StandardResultsSetPagination

    def get_queryset(self):
        qs = User.objects.exclude(role='admin').prefetch_related('registrations').order_by('-id')
        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(user_code__icontains=search)
            )

        role = self.request.query_params.get('role', '').strip()
        if role in ['security', 'vehicle_owner']:
            qs = qs.filter(role=role)

        registrant_type = self.request.query_params.get('registrant_type', '').strip()
        if registrant_type in ['student', 'employee', 'fetcher']:
            qs = qs.filter(registrations__registrant_type=registrant_type).distinct()

        status_param = self.request.query_params.get('status', '').strip()
        if status_param == 'active':
            qs = qs.filter(is_active=True)
        elif status_param == 'disabled':
            qs = qs.filter(is_active=False)

        return qs


class UserDetailView(generics.RetrieveAPIView):
    """Get a single user by ID."""
    queryset           = User.objects.all()
    serializer_class   = UserSerializer
    permission_classes = [IsAdminRole]


class UserUpdateView(generics.UpdateAPIView):
    """Edit user details (full_name, email, role, photo)."""
    queryset           = User.objects.all()
    serializer_class   = UserUpdateSerializer
    permission_classes = [IsAdminRole]

    def get_serializer_context(self):
        return {**super().get_serializer_context(), 'request': self.request}

    def perform_update(self, serializer):
        old_user = serializer.instance
        changes = []
        for field in ['full_name', 'email', 'role']:
            old_val = getattr(old_user, field)
            new_val = serializer.validated_data.get(field, old_val)
            if old_val != new_val:
                changes.append(f"{field}: '{old_val}' → '{new_val}'")
        if 'photo' in serializer.validated_data:
            changes.append('photo updated')

        user = serializer.save()
        log_action(self.request, AuditLog.Action.USER_UPDATED, target_user=user, details='; '.join(changes))


class UserDeleteView(generics.DestroyAPIView):
    """Hard-delete a user, along with data that belongs to them (vehicles,
    registrations). Records that merely reference the user as an actor
    (audit logs, issued violations, scans, created events/notices) are left
    intact for accountability history — their FK just goes null."""
    queryset           = User.objects.all()
    serializer_class   = UserSerializer
    permission_classes = [IsAdminRole]

    def destroy(self, request, *args, **kwargs):
        from vehicles.models import Vehicle, VehicleRegistration

        user = self.get_object()
        if user.role == 'admin':
            return Response(
                {'detail': 'Cannot delete an admin from this endpoint.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        log_action(request, AuditLog.Action.USER_DELETED, target_user=user)
        Vehicle.objects.filter(user=user).delete()
        VehicleRegistration.objects.filter(user=user).delete()
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class UserToggleStatusView(APIView):
    """Toggle a user's is_active flag."""
    permission_classes = [IsAdminRole]

    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        if user.role == 'admin':
            return Response(
                {'detail': 'Cannot disable an admin account.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        old_status = user.is_active
        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])
        action = AuditLog.Action.USER_ENABLED if user.is_active else AuditLog.Action.USER_DISABLED
        log_action(request, action, target_user=user)
        return Response(UserSerializer(user).data)


class AdminReplaceView(APIView):
    """Create a new admin and delete the current admin."""
    permission_classes = [IsAdminRole]

    def post(self, request):
        serializer = AdminReplaceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        old_admin = request.user
        new_admin = serializer.save()
        log_action(request, AuditLog.Action.ADMIN_REPLACED, target_user=new_admin, details=f"Replaced admin: {old_admin.email}")
        old_admin.delete()
        return Response(
            {
                'detail': 'Admin replaced successfully.',
                'user': UserSerializer(new_admin).data,
            },
            status=status.HTTP_201_CREATED,
        )


class AdminCreateGuardView(APIView):
    """Admin creates a security-guard account with email + password credentials.
    Guards log in at the dedicated guard gate login page (credentials or QR badge)."""
    permission_classes = [IsAdminRole]

    def post(self, request):
        serializer = GuardCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        guard = serializer.save()
        log_action(request, AuditLog.Action.USER_CREATED, target_user=guard,
                   details=f'Guard account created: {guard.full_name}')
        return Response(UserSerializer(guard, context={'request': request}).data,
                        status=status.HTTP_201_CREATED)


class AdminCreateOwnerView(APIView):
    """Admin creates a vehicle-owner account directly.  Password is auto-generated
    and emailed to the owner; must_change_password is set."""
    permission_classes = [IsAdminRole]

    def post(self, request):
        serializer = AdminOwnerCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        owner = serializer.save()
        log_action(request, AuditLog.Action.USER_CREATED, target_user=owner,
                   details=f'Vehicle-owner account created by admin: {owner.full_name}')
        return Response(UserSerializer(owner, context={'request': request}).data,
                        status=status.HTTP_201_CREATED)


class DashboardStatsView(APIView):
    """Return dashboard stats. Admin gets full overview; security gets personal scan stats."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Count
        from django.utils import timezone
        from datetime import timedelta
        from scanning.models import AccessLog

        user = request.user
        today = timezone.localdate()
        week_ago = today - timedelta(days=7)

        if user.role == 'admin':
            from django.db.models.functions import ExtractWeekDay
            from vehicles.models import Vehicle, VehicleRegistration

            total_users         = User.objects.count()
            security_count      = User.objects.filter(role='security').count()
            vehicle_owner_count = User.objects.filter(role='vehicle_owner').count()
            active_users        = User.objects.filter(is_active=True).count()
            disabled_users      = User.objects.filter(is_active=False).count()

            total_vehicles        = Vehicle.objects.count()
            authorized_vehicles   = Vehicle.objects.filter(is_authorized=True).count()
            unauthorized_vehicles = total_vehicles - authorized_vehicles

            # Registration outcomes (pending / accepted / rejected) for the status pie
            reg_by_status = {
                row['status']: row['count']
                for row in VehicleRegistration.objects.values('status').annotate(count=Count('id'))
            }
            pending_registrations  = reg_by_status.get(VehicleRegistration.Status.PENDING, 0)
            accepted_registrations = reg_by_status.get(VehicleRegistration.Status.ACCEPTED, 0)
            rejected_registrations = reg_by_status.get(VehicleRegistration.Status.REJECTED, 0)

            # Vehicle-owner category breakdown (active owners split by owner_type,
            # plus a disabled slice) for the owners pie. Active-by-type + disabled
            # are mutually exclusive so they sum cleanly in a donut.
            owner_qs = User.objects.filter(role='vehicle_owner')
            owners_active_by_type = {
                row['owner_type']: row['count']
                for row in owner_qs.filter(is_active=True)
                                   .values('owner_type').annotate(count=Count('id'))
            }
            owners_disabled = owner_qs.filter(is_active=False).count()

            # Suppliers are their own model (not User owners) but form a registered
            # vehicle category alongside students/employees/fetchers. Counted by
            # plate (supplier vehicles), scoped to active supplier companies.
            from vehicles.models import SupplierPlate
            active_suppliers = SupplierPlate.objects.filter(supplier__is_active=True).count()

            # One aggregate query gives the full per-status picture for today
            today_by_status = {
                row['status']: row['count']
                for row in AccessLog.objects.filter(scanned_at__date=today)
                                            .values('status').annotate(count=Count('id'))
            }
            today_scans      = sum(today_by_status.values())
            week_scans       = AccessLog.objects.filter(scanned_at__date__gte=week_ago).count()
            authorized_today = today_by_status.get('authorized', 0)
            denied_today     = today_by_status.get('denied', 0) + today_by_status.get('wrong_day', 0)
            unknown_today    = today_by_status.get('unknown', 0)

            # Split today's authorized entries into registered owners vs visitor-pass
            # entries (visitor passes create an ownerless vehicle, so user is null).
            visitor_today    = AccessLog.objects.filter(
                scanned_at__date=today, status='authorized', vehicle__user__isnull=True,
            ).count()
            registered_today = authorized_today - visitor_today

            # Day distribution: authorized entries per weekday (Mon–Sat)
            # Django ExtractWeekDay: 1=Sunday, 2=Monday, …, 7=Saturday
            DAY_MAP = {2: 'Mon', 3: 'Tue', 4: 'Wed', 5: 'Thu', 6: 'Fri', 7: 'Sat'}
            day_rows = (
                AccessLog.objects
                .filter(status='authorized', scanned_at__date__gte=week_ago)
                .annotate(wd=ExtractWeekDay('scanned_at'))
                .filter(wd__in=DAY_MAP.keys())
                .values('wd')
                .annotate(count=Count('id'))
            )
            day_dist_map = {row['wd']: row['count'] for row in day_rows}
            day_distribution = [
                {'day': DAY_MAP[wd], 'count': day_dist_map.get(wd, 0)}
                for wd in sorted(DAY_MAP.keys())
            ]
            authorized_week = sum(day_dist_map.values())

            # Per-day registration load (Mon–Sat): how many active registrations
            # include each campus day, split accepted/pending, against the daily
            # slot capacity used by the public registration form.
            from vehicles.views import SCHEDULE_SLOT_LIMIT
            WEEK_DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
            day_registrations = []
            for _day in WEEK_DAYS:
                _day_qs = VehicleRegistration.objects.filter(campus_days__contains=[_day])
                day_registrations.append({
                    'day':      _day,
                    'accepted': _day_qs.filter(status=VehicleRegistration.Status.ACCEPTED).count(),
                    'pending':  _day_qs.filter(status=VehicleRegistration.Status.PENDING).count(),
                    'capacity': SCHEDULE_SLOT_LIMIT,
                })

            # Violations & visitor passes — surfaced on the dashboard KPI strip
            from violations.models import Violation
            from scanning.models import VisitorPass
            open_violations = Violation.objects.filter(is_resolved=False).count()
            fee_imposed     = Violation.objects.filter(status=Violation.Status.FEE_IMPOSED).count()
            active_passes   = VisitorPass.objects.filter(
                valid_date=today, status=VisitorPass.Status.ACTIVE
            ).count()

            recent_admin_logs    = AuditLog.objects.select_related('actor', 'target_user').filter(
                actor__role='admin'
            ).order_by('-created_at')[:10]
            recent_security_logs = AuditLog.objects.select_related('actor', 'target_user').filter(
                actor__role='security'
            ).order_by('-created_at')[:10]

            data = {
                'role': 'admin',
                'users': {
                    'total':         total_users,
                    'security':      security_count,
                    'vehicle_owner': vehicle_owner_count,
                    'active':        active_users,
                    'disabled':      disabled_users,
                },
                'vehicles': {
                    'total':        total_vehicles,
                    'authorized':   authorized_vehicles,
                    'unauthorized': unauthorized_vehicles,
                },
                'registrations': {
                    'pending':  pending_registrations,
                    'accepted': accepted_registrations,
                    'rejected': rejected_registrations,
                    'total':    pending_registrations + accepted_registrations + rejected_registrations,
                },
                'owners': {
                    'student':  owners_active_by_type.get('student', 0),
                    'employee': owners_active_by_type.get('employee', 0),
                    'fetcher':  owners_active_by_type.get('fetcher', 0),
                    'visitor':  owners_active_by_type.get('visitor', 0),
                    'disabled': owners_disabled,
                    'total':    vehicle_owner_count,
                },
                'suppliers': {
                    'active': active_suppliers,
                },
                'scans': {
                    'today':            today_scans,
                    'week':             week_scans,
                    'authorized_today': authorized_today,
                    'authorized_week':  authorized_week,
                    'registered_today': registered_today,
                    'visitor_today':    visitor_today,
                    'denied_today':     denied_today,
                    'unknown_today':    unknown_today,
                    'today_by_status':  today_by_status,
                },
                'violations': {
                    'open':        open_violations,
                    'fee_imposed': fee_imposed,
                },
                'visitor_passes': {
                    'active_today': active_passes,
                },
                'day_distribution': day_distribution,
                'day_registrations': day_registrations,
                'recent_activity': {
                    'admin':    AuditLogSerializer(recent_admin_logs, many=True).data,
                    'security': AuditLogSerializer(recent_security_logs, many=True).data,
                },
            }
        else:
            my_scans_today = AccessLog.objects.filter(scanned_by=user, scanned_at__date=today).count()
            my_scans_week  = AccessLog.objects.filter(scanned_at__date__gte=week_ago, scanned_by=user).count()
            my_total_scans = AccessLog.objects.filter(scanned_by=user).count()
            my_authorized  = AccessLog.objects.filter(scanned_by=user, scanned_at__date=today, status='authorized').count()
            my_denied      = AccessLog.objects.filter(
                scanned_by=user, scanned_at__date=today, status__in=['denied', 'wrong_day']
            ).count()

            my_access_logs = AccessLog.objects.filter(scanned_by=user).order_by('-scanned_at')[:10]

            from scanning.serializers import AccessLogSerializer
            data = {
                'role': 'security',
                'scans': {
                    'today':            my_scans_today,
                    'week':             my_scans_week,
                    'total':            my_total_scans,
                    'authorized_today': my_authorized,
                    'denied_today':     my_denied,
                },
                'recent_scans': AccessLogSerializer(my_access_logs, many=True).data,
            }

        return Response(data)


# ──────────────────────────────────────────────
#  Audit Log Views
# ──────────────────────────────────────────────

class AuditLogListView(generics.ListAPIView):
    """List audit logs - admin only."""
    serializer_class   = AuditLogSerializer
    permission_classes = [IsAdminRole]
    pagination_class   = StandardResultsSetPagination

    def get_queryset(self):
        qs = AuditLog.objects.select_related('actor', 'target_user').all()

        action = self.request.query_params.get('action', '').strip()
        if action:
            qs = qs.filter(action=action)

        date_from = self.request.query_params.get('date_from', '').strip()
        date_to   = self.request.query_params.get('date_to', '').strip()
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Search matches the actor (code / name / email) or the details text
        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(actor__user_code__icontains=search) |
                Q(actor__full_name__icontains=search) |
                Q(actor__email__icontains=search) |
                Q(details__icontains=search)
            )

        return qs.order_by('-created_at')


# Brand colour + logo shared by the Audit Log Excel/PDF reports.
REPORT_BRAND_HEX = '2A2B61'
REPORT_LOGO_PATH = os.path.join(settings.BASE_DIR, 'report_assets', 'slclogo.jpg')


def _filter_audit_logs(request):
    """Apply the same filters the Audit Log UI uses.

    Returns (ordered_queryset, filters_desc) so the Excel and PDF exports stay
    identical to what the operator sees on screen.
    """
    qs = AuditLog.objects.select_related('actor', 'target_user').all()
    action    = request.query_params.get('action', '').strip()
    date_from = request.query_params.get('date_from', '').strip()
    date_to   = request.query_params.get('date_to', '').strip()
    search    = request.query_params.get('search', '').strip()
    if action:
        qs = qs.filter(action=action)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if search:
        qs = qs.filter(
            Q(actor__user_code__icontains=search) |
            Q(actor__full_name__icontains=search) |
            Q(actor__email__icontains=search) |
            Q(details__icontains=search)
        )

    action_labels = dict(AuditLog.Action.choices)
    filters_desc = []
    if action:
        filters_desc.append(f"Action: {action_labels.get(action, action)}")
    if date_from or date_to:
        filters_desc.append(f"Period: {date_from or 'start'} to {date_to or 'today'}")
    if search:
        filters_desc.append(f"Search: '{search}'")
    return qs.order_by('-created_at'), filters_desc


class AuditLogExportView(APIView):
    """Download the (filtered) audit log as a branded Excel report — admin only."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        from django.http import HttpResponse
        from django.utils import timezone as tz
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.drawing.image import Image as XLImage

        qs, filters_desc = _filter_audit_logs(request)
        rows = qs[:5000]
        action_labels = dict(AuditLog.Action.choices)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Vehicle Log'

        # ── Branded header band: logo (col A) + title/subtitle (cols C–F) ──
        if os.path.exists(REPORT_LOGO_PATH):
            try:
                logo = XLImage(REPORT_LOGO_PATH)
                logo.width, logo.height = 150, 55
                ws.add_image(logo, 'A1')
            except Exception:
                pass  # never let a logo hiccup break the export
        for rh in (1, 2, 3):
            ws.row_dimensions[rh].height = 18

        ws.merge_cells('C1:F1')
        ws['C1'] = 'Saint Louis College — Vehicle Management System · Vehicle Log Report'
        ws['C1'].font = Font(bold=True, size=13, color=REPORT_BRAND_HEX)

        ws.merge_cells('C2:F2')
        ws['C2'] = (
            f"Generated {tz.localtime().strftime('%B %d, %Y %I:%M %p')} "
            f"by {getattr(request.user, 'full_name', '')} · "
            + ('; '.join(filters_desc) if filters_desc else 'All records')
            + f" · {rows.count() if hasattr(rows, 'count') else len(rows)} entries"
        )
        ws['C2'].font = Font(size=10, color='666666')

        # ── Column headers (row 4) ──
        headers = ['#', 'Date & Time', 'Actor', 'Role', 'Action', 'Details']
        header_fill = PatternFill('solid', fgColor=REPORT_BRAND_HEX)
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=title)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical='center')

        for width, col in zip([5, 21, 24, 12, 20, 95], 'ABCDEF'):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = 'A5'

        wrap = Alignment(wrap_text=True, vertical='top')
        last_row = 4
        for i, log in enumerate(rows, start=1):
            r = 4 + i
            last_row = r
            actor = log.actor.full_name if log.actor else 'System'
            role  = (log.actor.role if log.actor else '').replace('_', ' ').title()
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2,
                    value=tz.localtime(log.created_at).strftime('%b %d, %Y %I:%M:%S %p'))
            ws.cell(row=r, column=3, value=actor)
            ws.cell(row=r, column=4, value=role)
            ws.cell(row=r, column=5, value=action_labels.get(log.action, log.action))
            c = ws.cell(row=r, column=6, value=log.details or '')
            c.alignment = wrap

        # ── Footer ──
        foot_row = last_row + 2
        ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=6)
        fcell = ws.cell(row=foot_row, column=1,
                        value='— End of report · Saint Louis College CDSO · Confidential —')
        fcell.font = Font(size=9, italic=True, color='999999')
        fcell.alignment = Alignment(horizontal='center')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        stamp = tz.localtime().strftime('%Y%m%d-%H%M')
        response['Content-Disposition'] = f'attachment; filename="vehicle-log-report-{stamp}.xlsx"'
        wb.save(response)
        return response


class AuditLogPdfExportView(APIView):
    """Download the (filtered) audit log as a branded PDF report — admin only."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        from django.http import HttpResponse
        from django.utils import timezone as tz
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage,
        )

        qs, filters_desc = _filter_audit_logs(request)
        rows = list(qs[:5000])
        action_labels = dict(AuditLog.Action.choices)

        brand = colors.HexColor(f'#{REPORT_BRAND_HEX}')
        generated_by = getattr(request.user, 'full_name', '')
        generated_at = tz.localtime().strftime('%B %d, %Y %I:%M %p')
        subtitle = (
            ('; '.join(filters_desc) if filters_desc else 'All records')
            + f" · {len(rows)} entries"
        )

        # ── Header/footer painted on every page ──
        def draw_frame(canvas, doc):
            canvas.saveState()
            w, h = landscape(A4)
            # Logo top-left
            if os.path.exists(REPORT_LOGO_PATH):
                try:
                    canvas.drawImage(
                        REPORT_LOGO_PATH, 15 * mm, h - 26 * mm,
                        width=30 * mm, height=16 * mm, preserveAspectRatio=True, mask='auto',
                    )
                except Exception:
                    pass
            canvas.setFillColor(brand)
            canvas.setFont('Helvetica-Bold', 13)
            canvas.drawString(50 * mm, h - 16 * mm, 'Saint Louis College — Vehicle Management System')
            canvas.setFont('Helvetica', 10)
            canvas.setFillColor(colors.HexColor('#666666'))
            canvas.drawString(50 * mm, h - 21 * mm, 'Vehicle Log Report')
            # Header rule
            canvas.setStrokeColor(brand)
            canvas.setLineWidth(1)
            canvas.line(15 * mm, h - 28 * mm, w - 15 * mm, h - 28 * mm)
            # Footer
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.HexColor('#999999'))
            canvas.drawString(15 * mm, 10 * mm,
                              f'Generated {generated_at} by {generated_by} · Saint Louis College CDSO · Confidential')
            canvas.drawRightString(w - 15 * mm, 10 * mm, f'Page {doc.page}')
            canvas.restoreState()

        buf_response = HttpResponse(content_type='application/pdf')
        stamp = tz.localtime().strftime('%Y%m%d-%H%M')
        buf_response['Content-Disposition'] = f'attachment; filename="vehicle-log-report-{stamp}.pdf"'

        doc = SimpleDocTemplate(
            buf_response, pagesize=landscape(A4),
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=32 * mm, bottomMargin=16 * mm,
            title='Vehicle Log Report',
        )

        cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=10)
        head_style = ParagraphStyle('head', fontName='Helvetica-Bold', fontSize=9,
                                    textColor=colors.white, leading=11)

        sub_style = ParagraphStyle('sub', fontName='Helvetica', fontSize=9,
                                   textColor=colors.HexColor('#666666'))
        story = [Paragraph(subtitle, sub_style), Spacer(1, 4 * mm)]

        table_head = [Paragraph(t, head_style) for t in
                      ['#', 'Date &amp; Time', 'Actor', 'Role', 'Action', 'Details']]
        data = [table_head]
        for i, log in enumerate(rows, start=1):
            actor = log.actor.full_name if log.actor else 'System'
            role  = (log.actor.role if log.actor else '').replace('_', ' ').title()
            details = (log.details or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            data.append([
                Paragraph(str(i), cell_style),
                Paragraph(tz.localtime(log.created_at).strftime('%b %d, %Y %I:%M:%S %p'), cell_style),
                Paragraph(actor, cell_style),
                Paragraph(role, cell_style),
                Paragraph(action_labels.get(log.action, log.action), cell_style),
                Paragraph(details, cell_style),
            ])

        if len(data) == 1:
            data.append([Paragraph('No records match the selected filters.', cell_style)]
                        + [Paragraph('', cell_style) for _ in range(5)])

        col_widths = [10 * mm, 34 * mm, 42 * mm, 22 * mm, 38 * mm, 121 * mm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F5FA')]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E8F0')),
        ]))
        story.append(table)

        doc.build(story, onFirstPage=draw_frame, onLaterPages=draw_frame)
        return buf_response


# ── System Backup & Restore ─────────────────────────────────────────────────
# App labels whose data is captured in a backup. Excludes contenttypes,
# permissions, sessions, admin log entries and token blacklists (volatile /
# rebuildable) — the schema itself is versioned by migrations.
BACKUP_APPS = ['accounts', 'vehicles', 'scanning', 'violations', 'realtime']

# High-volume ML artefacts (plate-recognition crops and training samples) are
# rebuildable and would bloat every backup by ~20k rows / many MB, making
# restore impractically slow. Business data — including the access log — is kept.
BACKUP_EXCLUDE = ['scanning.platerecognitionrecord', 'scanning.mltrainingsample']


class SystemBackupView(APIView):
    """Download a JSON snapshot of all application data — admin (CDSO) only."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        import io
        from django.core.management import call_command
        from django.http import HttpResponse
        from django.utils import timezone as tz

        buf = io.StringIO()
        call_command('dumpdata', *BACKUP_APPS, exclude=BACKUP_EXCLUDE, indent=2, stdout=buf)

        response = HttpResponse(buf.getvalue(), content_type='application/json')
        stamp = tz.localtime().strftime('%Y%m%d-%H%M')
        response['Content-Disposition'] = f'attachment; filename="slc-vms-backup-{stamp}.json"'
        log_action(request, AuditLog.Action.RECORD_UPDATED, details='System backup downloaded')
        return response


class SystemRestoreView(APIView):
    """Restore application data from an uploaded JSON backup — admin (CDSO) only.

    Safety measures: admin-only, file validated as a JSON fixture, an automatic
    pre-restore snapshot of current data is saved to disk, and the load runs in
    a single transaction that rolls back completely on any error.
    """
    permission_classes = [IsAdminRole]

    def post(self, request):
        import io, json, os, tempfile
        from django.core.management import call_command
        from django.db import transaction
        from django.utils import timezone as tz

        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'No backup file provided.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if upload.size > 50 * 1024 * 1024:
            return Response({'error': 'Backup file is too large (max 50 MB).'},
                            status=status.HTTP_400_BAD_REQUEST)

        raw = upload.read()
        try:
            parsed = json.loads(raw.decode('utf-8'))
            if not isinstance(parsed, list):
                raise ValueError('not a fixture list')
        except Exception:
            return Response({'error': 'Invalid backup file — expected a JSON data fixture.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # 1) Auto safety snapshot of current data before overwriting anything.
        safety = io.StringIO()
        call_command('dumpdata', *BACKUP_APPS, exclude=BACKUP_EXCLUDE, indent=2, stdout=safety)
        safety_dir = os.path.join(settings.BASE_DIR, 'backups')
        os.makedirs(safety_dir, exist_ok=True)
        safety_name = f"pre-restore-{tz.localtime().strftime('%Y%m%d-%H%M%S')}.json"
        with open(os.path.join(safety_dir, safety_name), 'w', encoding='utf-8') as fh:
            fh.write(safety.getvalue())

        # 2) Load the uploaded fixture atomically (rolls back on any error).
        tmp = tempfile.NamedTemporaryFile('wb', suffix='.json', delete=False)
        try:
            tmp.write(raw)
            tmp.close()
            with transaction.atomic():
                call_command('loaddata', tmp.name, verbosity=0)
        except Exception as exc:
            return Response(
                {'error': f'Restore failed and was rolled back. No changes were applied. ({exc})',
                 'safety_backup': safety_name},
                status=status.HTTP_400_BAD_REQUEST,
            )
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

        log_action(request, AuditLog.Action.RECORD_UPDATED,
                   details=f'System restore from backup ({len(parsed)} records)')
        return Response({'restored': len(parsed), 'safety_backup': safety_name},
                        status=status.HTTP_200_OK)


class AuditLogClearView(APIView):
    """Delete all audit log records — admin only."""
    permission_classes = [IsAdminRole]

    def delete(self, request):
        deleted_count, _ = AuditLog.objects.all().delete()
        return Response({'deleted': deleted_count}, status=status.HTTP_200_OK)


class AuditLogStatsView(APIView):
    """Get audit log statistics."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        from django.db.models import Count
        from django.utils import timezone
        from datetime import timedelta
        
        today = timezone.localdate()
        week_ago = today - timedelta(days=7)
        
        stats = {
            'total_logs': AuditLog.objects.count(),
            'today_logs': AuditLog.objects.filter(created_at__date=today).count(),
            'week_logs': AuditLog.objects.filter(created_at__date__gte=week_ago).count(),
            'by_action': AuditLog.objects.values('action').annotate(count=Count('action')),
        }
        return Response(stats)


# ──────────────────────────────────────────────
#  Password Change (any authenticated user)
# ──────────────────────────────────────────────

class ChangePasswordView(APIView):
    """Allow any authenticated user to change their own password.
    Clears the must_change_password flag after a successful change."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        current_password = request.data.get('current_password', '').strip()
        new_password = request.data.get('new_password', '').strip()
        confirm_password = request.data.get('confirm_password', '').strip()

        if not current_password:
            return Response({'error': 'Current password is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not user.check_password(current_password):
            return Response({'error': 'Current password is incorrect.'}, status=status.HTTP_400_BAD_REQUEST)
        if not new_password:
            return Response({'error': 'New password is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if new_password != confirm_password:
            return Response({'error': 'New passwords do not match.'}, status=status.HTTP_400_BAD_REQUEST)
        if current_password == new_password:
            return Response({'error': 'New password must be different from current password.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate strength
        import re
        errors = []
        if len(new_password) < 8:
            errors.append('Password must be at least 8 characters.')
        if not re.search(r'[A-Z]', new_password):
            errors.append('Password must contain at least one uppercase letter.')
        if not re.search(r'[a-z]', new_password):
            errors.append('Password must contain at least one lowercase letter.')
        if not re.search(r'[0-9]', new_password):
            errors.append('Password must contain at least one number.')
        if not re.search(r'[!@#$%^&*()_+\-=\[\]{};\'\"\\|,.<>\/?]', new_password):
            errors.append('Password must contain at least one special character.')
        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.must_change_password = False
        user.save(update_fields=['password', 'must_change_password'])
        return Response({'message': 'Password changed successfully.'})


# ──────────────────────────────────────────────
#  Vehicle Owner: own registration record
# ──────────────────────────────────────────────

class MyRegistrationView(APIView):
    """Returns the VehicleRegistration record for the logged-in vehicle owner."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'vehicle_owner':
            return Response({'error': 'Only vehicle owners can access this endpoint.'}, status=status.HTTP_403_FORBIDDEN)
        from vehicles.models import VehicleRegistration
        from vehicles.serializers import VehicleRegistrationSerializer
        registration = (
            VehicleRegistration.objects
            .filter(Q(user=request.user) | Q(email=request.user.email), status='accepted')
            .order_by('-reviewed_at')
            .first()
        )
        if not registration:
            return Response({'error': 'No accepted registration found for this account.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(VehicleRegistrationSerializer(registration).data)


# ──────────────────────────────────────────────
#  Guard QR Login (passwordless, for gate stations)
# ──────────────────────────────────────────────

class GuardQrLoginView(APIView):
    """
    Authenticate a security guard by scanning their QR badge.
    QR format: SLC-GUARD:{user_code}:{guard_qr_secret}
    Logs out any previously active guard session implicitly — the new JWT supersedes the old one.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        import uuid as _uuid
        from rest_framework_simplejwt.tokens import RefreshToken

        qr_data = (request.data.get('qr_data') or '').strip()
        if not qr_data.startswith('SLC-GUARD:'):
            return Response({'detail': 'Invalid QR code.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            _, user_code, secret_str = qr_data.split(':', 2)
            secret = _uuid.UUID(secret_str)
        except (ValueError, AttributeError):
            return Response({'detail': 'Malformed QR code.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(user_code=user_code, guard_qr_secret=secret, role='security', is_active=True)
        except User.DoesNotExist:
            return Response({'detail': 'QR code not recognised or guard account is disabled.'}, status=status.HTTP_401_UNAUTHORIZED)

        # QR login is passwordless — refuse it until the guard has completed
        # their first credentials login and replaced the temporary password.
        if user.must_change_password:
            return Response(
                {'detail': 'QR login is disabled until you sign in with your credentials and change your temporary password.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        refresh = RefreshToken.for_user(user)
        refresh['role'] = user.role
        refresh['full_name'] = user.full_name
        refresh['email'] = user.email
        refresh['must_change_password'] = user.must_change_password

        AuditLog.objects.create(
            actor=user,
            action=AuditLog.Action.SCAN,
            details=f'Guard QR login: {user.full_name} ({user.user_code})',
            ip_address=get_client_ip(request),
        )

        return Response({
            'access':  str(refresh.access_token),
            'refresh': str(refresh),
            'user': {
                'id':                 user.id,
                'user_code':          user.user_code,
                'full_name':          user.full_name,
                'email':              user.email,
                'role':               user.role,
                'must_change_password': user.must_change_password,
                'photo_url': request.build_absolute_uri(user.photo.url) if user.photo else None,
            },
        })


class GuardQrCodeView(APIView):
    """
    Generate (or retrieve) a guard's QR secret.
    Admin: GET /accounts/guard-qr/{pk}/ — returns the QR payload string for that guard.
    The guard themselves can also hit this endpoint to get their own code.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        import uuid as _uuid

        if request.user.role == 'admin':
            user = get_object_or_404(User, pk=pk, role='security')
        elif request.user.pk == int(pk) and request.user.role == 'security':
            user = request.user
        else:
            return Response({'detail': 'Not authorised.'}, status=status.HTTP_403_FORBIDDEN)

        if user.must_change_password:
            return Response(
                {'detail': 'QR badge is locked — this guard must log in with their credentials and change their temporary password first.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if not user.guard_qr_secret:
            user.guard_qr_secret = _uuid.uuid4()
            User.objects.filter(pk=user.pk).update(guard_qr_secret=user.guard_qr_secret)

        qr_payload = f'SLC-GUARD:{user.user_code}:{user.guard_qr_secret}'
        return Response({
            'user_code':   user.user_code,
            'full_name':   user.full_name,
            'qr_payload':  qr_payload,
            'photo_url':   request.build_absolute_uri(user.photo.url) if user.photo else None,
        })


class PasswordResetRequestView(APIView):
    """Step 1: accept an email, generate a token, send a reset link."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_encode
        from django.utils.encoding import force_bytes
        from django.core.mail import send_mail
        from django.conf import settings as django_settings

        email = request.data.get('email', '').strip().lower()
        if not email:
            return Response({'error': 'Email is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Always return the same message to avoid leaking which emails exist
        SAFE_MSG = 'If an account with that email exists, a password reset link has been sent.'

        try:
            user = User.objects.get(email__iexact=email, is_active=True)
        except User.DoesNotExist:
            return Response({'message': SAFE_MSG})

        token = default_token_generator.make_token(user)
        uid   = urlsafe_base64_encode(force_bytes(user.pk))
        frontend_url = getattr(django_settings, 'FRONTEND_URL', 'http://localhost:5173')
        reset_link   = f"{frontend_url}/reset-password?uid={uid}&token={token}"

        html_message = f"""
        <html>
          <body style="font-family:Arial,sans-serif;color:#1A1D2E;background:#F0F2F7;padding:20px;margin:0;">
            <div style="max-width:540px;margin:0 auto;background:#fff;border-radius:12px;border-top:4px solid #2A2B61;box-shadow:0 4px 20px rgba(0,0,0,.08);overflow:hidden;">
              <div style="padding:28px 32px 24px;">
                <h2 style="color:#2A2B61;margin:0 0 8px;">Password Reset Request</h2>
                <p style="color:#5A5F72;font-size:14px;margin:0 0 20px;">
                  We received a request to reset the password for your SLC Vehicle Management System account.
                </p>
                <p style="margin:0 0 8px;">Hello, <strong>{user.full_name or user.email}</strong>,</p>
                <p style="color:#5A5F72;font-size:14px;margin:0 0 24px;">
                  Click the button below to set a new password. This link expires in <strong>1 hour</strong>.
                </p>
                <div style="text-align:center;margin:0 0 24px;">
                  <a href="{reset_link}"
                     style="display:inline-block;padding:13px 32px;background:#2A2B61;color:#fff;
                            border-radius:10px;font-size:15px;font-weight:600;text-decoration:none;">
                    Reset My Password
                  </a>
                </div>
                <p style="color:#9CA3B0;font-size:12px;margin:0 0 8px;">
                  If the button doesn't work, copy and paste this link into your browser:
                </p>
                <p style="word-break:break-all;font-size:12px;color:#2A2B61;margin:0 0 24px;">{reset_link}</p>
                <p style="color:#9CA3B0;font-size:12px;margin:0;">
                  If you did not request a password reset, you can safely ignore this email.
                  Your password will not change.
                </p>
              </div>
              <div style="background:#F8FAFC;border-top:1px solid #E2E6EE;padding:14px 32px;text-align:center;">
                <p style="font-size:12px;color:#7C80A3;margin:0;">Saint Louis College Vehicle Management System</p>
                <p style="font-size:11px;color:#B0B4C7;margin:4px 0 0;">This is an automated message. Please do not reply.</p>
              </div>
            </div>
          </body>
        </html>
        """
        
        send_mail(
            subject='SLC Vehicle Management — Password Reset',
            message=(
                f"Hello {user.full_name or user.email},\n\n"
                f"Reset your password by visiting:\n{reset_link}\n\n"
                f"This link expires in 1 hour.\n\n"
                f"If you did not request this, ignore this email."
            ),
            from_email=django_settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=True,
        )

        return Response({'message': SAFE_MSG})


class PasswordResetConfirmView(APIView):
    """Step 2: validate the token and set the new password."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        import re
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_decode
        from django.utils.encoding import force_str

        uid             = request.data.get('uid', '').strip()
        token           = request.data.get('token', '').strip()
        new_password    = request.data.get('new_password', '').strip()
        confirm_password = request.data.get('confirm_password', '').strip()

        if not all([uid, token, new_password, confirm_password]):
            return Response({'error': 'All fields are required.'}, status=status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return Response({'error': 'Passwords do not match.'}, status=status.HTTP_400_BAD_REQUEST)

        # Decode UID and fetch user
        try:
            pk   = force_str(urlsafe_base64_decode(uid))
            user = User.objects.get(pk=pk)
        except (User.DoesNotExist, ValueError, TypeError, Exception):
            return Response({'error': 'Invalid reset link.'}, status=status.HTTP_400_BAD_REQUEST)

        if not default_token_generator.check_token(user, token):
            return Response(
                {'error': 'This reset link has expired or is invalid. Please request a new one.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate password strength (same rules as ChangePasswordView)
        errors = []
        if len(new_password) < 8:
            errors.append('Password must be at least 8 characters.')
        if not re.search(r'[A-Z]', new_password):
            errors.append('Password must contain at least one uppercase letter.')
        if not re.search(r'[a-z]', new_password):
            errors.append('Password must contain at least one lowercase letter.')
        if not re.search(r'[0-9]', new_password):
            errors.append('Password must contain at least one number.')
        if not re.search(r'[!@#$%^&*()_+\-=\[\]{};\'\"\\|,.<>\/?]', new_password):
            errors.append('Password must contain at least one special character.')
        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.must_change_password = False
        user.save(update_fields=['password', 'must_change_password'])

        return Response({
            'message': 'Password reset successfully. You can now log in with your new password.',
            'role': user.role,
        })


# ──────────────────────────────────────────────
#  Guard QR Login & Shift Management
# ──────────────────────────────────────────────

class QRLoginView(APIView):
    """Guard scans their QR badge — clocks out previous shift, creates new shift, issues JWT."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from scanning.models import GuardShift
        from scanning.serializers import GuardShiftSerializer
        from rest_framework_simplejwt.tokens import RefreshToken

        token_str  = (request.data.get('qr_token') or '').strip()
        gate_param = (request.data.get('gate')     or '').strip()

        if not token_str:
            return Response({'error': 'qr_token is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            guard = User.objects.get(qr_token=token_str, role='security', is_active=True)
        except (User.DoesNotExist, ValueError):
            return Response({'error': 'Invalid or unrecognized QR code.'}, status=status.HTTP_401_UNAUTHORIZED)

        # QR login is passwordless — refuse it until the guard has completed
        # their first credentials login and replaced the temporary password.
        if guard.must_change_password:
            return Response(
                {'error': 'QR login is disabled until you sign in with your credentials and change your temporary password.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        from scanning.models import Gate
        valid_gates = Gate.active_ids()
        gate = gate_param if gate_param in valid_gates else guard.gate_assignment
        if not gate or gate not in valid_gates:
            return Response(
                {'error': 'Gate selection required. Please choose a gate before scanning.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Persist the shift gate on the guard's profile. Server-side scan
        # attribution (manual entry, override, exit, HTTP/WS scans) reads
        # request.user.gate_assignment to tag each log; without this it stays
        # None and those scans fall to the orphan 'main' gate, invisible in
        # every gate's Vehicle Log.
        if guard.gate_assignment != gate:
            User.objects.filter(pk=guard.pk).update(gate_assignment=gate)
            guard.gate_assignment = gate

        now = timezone.now()

        # Clock out whoever is currently active at this gate (the previous guard)
        GuardShift.objects.filter(gate=gate, clocked_out_at__isnull=True).update(
            clocked_out_at=now,
            clocked_out_by=guard,
        )

        shift   = GuardShift.objects.create(guard=guard, gate=gate)
        refresh = RefreshToken.for_user(guard)

        return Response({
            'access':  str(refresh.access_token),
            'refresh': str(refresh),
            'user': {
                'id':                   guard.id,
                'user_code':            guard.user_code,
                'full_name':            guard.full_name,
                'email':                guard.email,
                'role':                 guard.role,
                'gate_assignment':      gate,   # current shift gate (also persisted above)
                'must_change_password': guard.must_change_password,
            },
            'shift': GuardShiftSerializer(shift).data,
        })


class GuardCredentialLoginView(APIView):
    """Guard logs in with email + password at the gate station — clocks out the
    previous shift at the selected gate, creates a new shift, issues JWT.
    Only accounts with role='security' may use this endpoint; everyone else
    logs in through the regular /api/auth/login/ endpoint."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from scanning.models import GuardShift
        from scanning.serializers import GuardShiftSerializer
        from rest_framework_simplejwt.tokens import RefreshToken

        email      = (request.data.get('email')    or '').strip().lower()
        password   = request.data.get('password')  or ''
        gate_param = (request.data.get('gate')     or '').strip()

        if not email or not password:
            return Response({'error': 'Email and password are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            guard = User.objects.get(email__iexact=email, role='security')
        except User.DoesNotExist:
            return Response({'error': 'Incorrect email or password.'}, status=status.HTTP_401_UNAUTHORIZED)

        if not guard.check_password(password):
            return Response({'error': 'Incorrect email or password.'}, status=status.HTTP_401_UNAUTHORIZED)

        if not guard.is_active:
            return Response(
                {'error': 'Your account has been disabled. Please contact the administrator.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        from scanning.models import Gate
        valid_gates = Gate.active_ids()
        gate = gate_param if gate_param in valid_gates else guard.gate_assignment
        if not gate or gate not in valid_gates:
            return Response(
                {'error': 'Gate selection required. Please choose a gate before logging in.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Persist the shift gate on the guard's profile. Server-side scan
        # attribution (manual entry, override, exit, HTTP/WS scans) reads
        # request.user.gate_assignment to tag each log; without this it stays
        # None and those scans fall to the orphan 'main' gate, invisible in
        # every gate's Vehicle Log.
        if guard.gate_assignment != gate:
            User.objects.filter(pk=guard.pk).update(gate_assignment=gate)
            guard.gate_assignment = gate

        now = timezone.now()

        # Clock out whoever is currently active at this gate (the previous guard)
        GuardShift.objects.filter(gate=gate, clocked_out_at__isnull=True).update(
            clocked_out_at=now,
            clocked_out_by=guard,
        )

        shift   = GuardShift.objects.create(guard=guard, gate=gate)
        refresh = RefreshToken.for_user(guard)

        AuditLog.objects.create(
            actor=guard,
            action=AuditLog.Action.SCAN,
            details=f'Guard credential login: {guard.full_name} ({guard.user_code}) at {gate}',
            ip_address=get_client_ip(request),
        )

        return Response({
            'access':  str(refresh.access_token),
            'refresh': str(refresh),
            'user': {
                'id':                   guard.id,
                'user_code':            guard.user_code,
                'full_name':            guard.full_name,
                'email':                guard.email,
                'role':                 guard.role,
                'gate_assignment':      gate,   # current shift gate (also persisted above)
                'must_change_password': guard.must_change_password,
            },
            'shift': GuardShiftSerializer(shift).data,
        })


class GuardQrAvailabilityView(APIView):
    """Public: whether a guard can log in by QR badge (i.e. has completed
    their first credentials login and password change). With ?email= the
    check is for that specific guard; without it, whether any guard can.
    The gate login page shows the QR Badge tab only when this is true.
    Returns only a boolean — no user data is exposed."""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        email = (request.query_params.get('email') or '').strip()
        qs = User.objects.filter(role='security', is_active=True, must_change_password=False)
        if email:
            qs = qs.filter(email__iexact=email)
        return Response({'qr_available': qs.exists()})


class GuardQRView(APIView):
    """Admin only: return a guard's QR token for badge printing."""
    permission_classes = [IsAdminRole]

    def get(self, request, pk):
        guard = get_object_or_404(User, pk=pk, role='security')
        if guard.must_change_password:
            return Response(
                {'detail': 'QR badge is locked — this guard must log in with their credentials and change their temporary password first.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return Response({
            'id':       guard.id,
            'full_name': guard.full_name,
            'qr_token': str(guard.qr_token),
        })


class NotificationListView(APIView):
    """Admin/CDSO: notification-bell feed with unread count."""
    permission_classes = [IsAdminOrCdso]

    def get(self, request):
        qs = Notification.objects.all()
        if request.query_params.get('unread_only') in ('1', 'true'):
            qs = qs.filter(is_read=False)
        try:
            limit = min(int(request.query_params.get('limit', 30)), 100)
        except (TypeError, ValueError):
            limit = 30
        unread_count = Notification.objects.filter(is_read=False).count()
        return Response({
            'results':      NotificationSerializer(qs[:limit], many=True).data,
            'unread_count': unread_count,
        })


class NotificationMarkReadView(APIView):
    """Admin/CDSO: mark notifications read — {"ids": [...]} or {"all": true}."""
    permission_classes = [IsAdminOrCdso]

    def post(self, request):
        if request.data.get('all'):
            qs = Notification.objects.filter(is_read=False)
        else:
            ids = request.data.get('ids') or []
            if not isinstance(ids, list) or not ids:
                return Response({'error': 'Provide "ids" (list) or "all": true.'}, status=400)
            qs = Notification.objects.filter(pk__in=ids, is_read=False)
        updated = qs.update(is_read=True)
        # queryset.update() skips post_save, so tell open pages explicitly
        if updated:
            try:
                from realtime.broadcast import broadcast_change
                broadcast_change('notification', 'updated')
            except Exception:
                pass
        return Response({'updated': updated})


class NotificationClearView(APIView):
    """Admin/CDSO: clear (delete) notifications from the bell feed.
    Body: {"read_only": true} deletes only already-read items; otherwise all."""
    permission_classes = [IsAdminOrCdso]

    def post(self, request):
        qs = Notification.objects.all()
        if request.data.get('read_only'):
            qs = qs.filter(is_read=True)
        deleted, _ = qs.delete()
        if deleted:
            try:
                from realtime.broadcast import broadcast_change
                broadcast_change('notification', 'updated')
            except Exception:
                pass
        return Response({'deleted': deleted})



